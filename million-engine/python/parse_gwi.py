#!/usr/bin/env python3
"""
Parse GWI Erste/Mastercard Excel export into nested JSON:

{
  "Gen Z": {
    "austria": {
      "Music Genres": {
        "80s music": {
          "index": 108.2,
          "col_pct": 52.7
        }
      },
      "Personal Interests": {
        "Adventure / extreme sports": { ... }
      }
    }
  }
}

Hierarchy: audience → country → question category → answer → metrics

Usage:
  python parse_gwi.py
  python parse_gwi.py /path/to/Export.xlsx -o ../data/gwi.json
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path("/Users/macbook2021/Desktop/mastercard/Erste_Mastercard-Export.xlsx")
DEFAULT_OUT = ROOT / "data" / "gwi.json"

AUDIENCE_MAP = {
    "Affluent - Erste Markets": "Affluent",
    "Gen Z - Erste Markets": "Gen Z",
    "All Internet Users (Audience Size)": "All Internet Users",
}

HEADER_ROW = 14
DATA_START = 20
METRICS_PER_BLOCK = 5  # Universe, Responses, Column %, Row %, Index


def slug_country(raw: str) -> str:
    """'Austria (Country)' → 'austria'"""
    name = re.sub(r"\s*\(Country\)\s*$", "", str(raw), flags=re.I).strip()
    return name.lower()


def audience_key(sheet_name: str, base_cell: str | None) -> str:
    if sheet_name in AUDIENCE_MAP:
        return AUDIENCE_MAP[sheet_name]
    for title, key in AUDIENCE_MAP.items():
        if sheet_name.startswith(title[:20]) or title.startswith(sheet_name[:20]):
            return key
    if base_cell and str(base_cell).startswith("Base:"):
        raw = str(base_cell).replace("Base:", "").strip()
        if " - " in raw:
            return raw.split(" - ")[0].strip()
        return raw
    return sheet_name.split("(")[0].strip()


def clean_category(raw: str | None) -> str:
    """'Music Genres*' → 'Music Genres'"""
    text = (raw or "").strip()
    text = re.sub(r"\*+$", "", text).strip()
    return text or "Other"


def split_answer(answer_raw: str | None, category: str) -> str:
    """
    '80s music (Music Genres*)' → '80s music'
    Falls back to full string if no parenthetical match.
    """
    text = (answer_raw or "").strip()
    if not text:
        return "unknown"
    # strip trailing " (Category)" / " (Category*)"
    m = re.match(r"^(.*?)\s*\([^)]*\)\s*$", text)
    if m:
        return m.group(1).strip() or text
    # if answer already equals category, keep as-is
    if text == category:
        return text
    return text


def read_countries(ws) -> list[tuple[int, str]]:
    out = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, col).value
        if not val:
            continue
        text = str(val).strip()
        if text.lower() in {"name", "metric", "totals", ""}:
            continue
        if "(country)" in text.lower():
            out.append((col, slug_country(text)))
    return out


def parse_block_metrics(ws, start_row: int) -> dict[str, dict[int, float]]:
    metrics: dict[str, dict[int, float]] = {}
    for offset in range(METRICS_PER_BLOCK):
        r = start_row + offset
        name = ws.cell(r, 4).value  # column D = Metric
        if not name:
            continue
        name = str(name).strip()
        by_col: dict[int, float] = {}
        for col in range(5, ws.max_column + 1):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                by_col[col] = float(v)
        metrics[name] = by_col
    return metrics


def parse_sheet(ws, sheet_name: str) -> dict:
    base = ws.cell(7, 2).value
    audience = audience_key(sheet_name, str(base) if base else None)
    countries = read_countries(ws)
    if not countries:
        raise ValueError(f"No countries found on sheet: {sheet_name}")

    # audience → country → category → answer → metrics
    tree: dict[str, dict] = {code: {} for _, code in countries}
    last_category = "Other"

    row = DATA_START
    max_row = ws.max_row or DATA_START
    while row <= max_row:
        metric = ws.cell(row, 4).value
        answer_cell = ws.cell(row, 3).value
        question_cell = ws.cell(row, 2).value

        if str(metric or "").strip() != "Universe":
            row += 1
            continue

        if str(answer_cell or "").strip() == "Totals":
            row += METRICS_PER_BLOCK
            continue

        if question_cell:
            last_category = clean_category(str(question_cell))
        category = last_category
        answer = split_answer(str(answer_cell) if answer_cell else None, category)

        metrics = parse_block_metrics(ws, row)
        index_by_col = metrics.get("Index", {})
        colpct_by_col = metrics.get("Column %", {})
        universe_by_col = metrics.get("Universe", {})
        responses_by_col = metrics.get("Responses", {})
        rowpct_by_col = metrics.get("Row %", {})

        for col, country in countries:
            col_raw = colpct_by_col.get(col)
            entry = {
                "index": round(index_by_col[col], 1) if col in index_by_col else None,
                "col_pct": round(col_raw * 100, 1) if col_raw is not None else None,
            }
            if col in universe_by_col:
                entry["universe"] = int(round(universe_by_col[col]))
            if col in responses_by_col:
                entry["responses"] = int(round(responses_by_col[col]))
            if col in rowpct_by_col:
                entry["row_pct"] = round(rowpct_by_col[col] * 100, 1)

            if entry["index"] is None and entry["col_pct"] is None:
                continue

            tree[country].setdefault(category, {})[answer] = entry

        row += METRICS_PER_BLOCK

    return {audience: tree}


def parse_workbook(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    root: dict = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        part = parse_sheet(ws, sheet_name)
        for audience, countries in part.items():
            root[audience] = countries
    wb.close()
    return root


def main() -> None:
    ap = argparse.ArgumentParser(description="GWI Excel → nested JSON")
    ap.add_argument("xlsx", nargs="?", default=str(DEFAULT_XLSX))
    ap.add_argument("-o", "--output", default=str(DEFAULT_OUT))
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        raise SystemExit(f"File not found: {xlsx}")

    data = parse_workbook(xlsx)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Wrote {out}")
    for audience, countries in data.items():
        sample_country = next(iter(countries))
        cats = list(countries[sample_country].keys())
        n_answers = sum(len(a) for a in countries[sample_country].values())
        print(f"  {audience}: {len(countries)} countries, ~{len(cats)} categories, ~{n_answers} answers")
        print(f"       categories: {cats[:6]}{'...' if len(cats) > 6 else ''}")

    # demo filter path
    try:
        demo = data["Gen Z"]["serbia"]["Music Genres"]["80s music"]
        print("\nDemo path Gen Z → serbia → Music Genres → 80s music:")
        print(" ", demo)
        aff = data["Affluent"]["serbia"]["Music Genres"]["80s music"]
        print("Compare Affluent serbia same answer:")
        print(" ", aff)
    except KeyError as e:
        print("Demo path missing:", e)


if __name__ == "__main__":
    main()
